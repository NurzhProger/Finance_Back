from django.http import HttpResponse
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from dirs.models import *
from docs.models import *
from django.db import connection
from wsgiref.util import FileWrapper
import json, os, datetime, tempfile


from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl.worksheet.page import PageMargins, PrintPageSetup


import jpype
jpype.startJVM()
from asposecells.api import Workbook

current_directory = os.path.dirname(__file__)



def child_orgs(ids):
    q = f"""with  orgs as (SELECT * FROM public.dirs_parent_organizations where _parent_id in {tuple(ids)}),
                maxorgs as (select _organization_id, max(_date) as _date from orgs group by _organization_id),
                actorgs as (select * from orgs where (_organization_id, _date) in (select * from maxorgs))
                select _organization_id as org_id from actorgs
                where not _organization_id in {tuple(ids)}"""
    with connection.cursor() as cursor:
        cursor.execute(q)
        result = cursor.fetchall()
    return result




# ********************************************
# *******ПЕЧАТНЫЕ ФОРМЫ ПО ДОКУМЕНТАМ*********
# ********************************************
@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report2728(request):

    datastr = request.body
    data = json.loads(datastr)
    id = data['id']
    tip_rep = data['tip_rep']

    # Получаем текущий каталог скрипта
    relative_path = os.path.join(current_directory, "report_template", "report_2728.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""WITH zapros as (SELECT * FROM public.docs_izm_exp_{tip_rep}
                            WHERE _izm_exp_id={id}),
                    org as (SELECT * FROM public.dirs_organization
                            WHERE id in (select _organization_id from zapros)),
                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from zapros)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                            WHERE id in (select _spec_id from zapros)),
                    res as  (select 
                                    fg.code as fg_code, fg.name_rus as fg_name, 
                                    abp.code as abp_code, abp.name_rus as abp_name, 
                                    org.codeorg as org_code, org.name_rus as org_name,
                                    right(pr.code, 3) as pr_code, pr.name_rus as pr_name,
                                    right(ppr.code, 3) as ppr_code, ppr.name_rus as ppr_name,
                                    spec.code as spec_code, spec.name_rus as spec_name,
                                    sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12 as god,
                                    sm1,sm2,sm3,sm4,sm5,sm6,sm7,sm8,sm9,sm10,sm11,sm12
                                from zapros

                                left join fkr on
                                    zapros._fkr_id = fkr.id
                                left join fg on
                                    fkr._funcgroup = fg.id	
                                left join abp on
                                    fkr._abp = abp.id
                                left join org on
                                    zapros._organization_id = org.id
                                left join pr on
                                    fkr._program_id = pr.id
                                left join ppr on
                                    fkr._podprogram_id = ppr.id
                                left join spec on
                                    zapros._spec_id = spec.id)
                                    
                SELECT  fg_code, abp_code, org_code, pr_code, ppr_code, spec_code, 
                        '' as name,
                        sum(god), sum(sm1), sum(sm2), sum(sm3), sum(sm4), sum(sm5), sum(sm6), sum(sm7), sum(sm8), sum(sm9), sum(sm10), sum(sm11), sum(sm12),
                        max(fg_name), max(abp_name), max(org_name), max(pr_name), max(ppr_name), max(spec_name)
                FROM res 
                GROUP BY ROLLUP (fg_code, abp_code, org_code, pr_code, ppr_code, spec_code)
                order by fg_code NULLS FIRST, abp_code NULLS FIRST, org_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    startrow = 17
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):
            if i > 20:
                continue
            if i == 7:  # если это наименование
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[2] == None and row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                elif (row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
                elif (row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[24])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[25])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            # Каждую строку меняем цвет фона чтобы было удобно пользователю
            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            # Если это группировка, то в следующем, меняем цвет, чтобы не было видно  (типа иерархия)
            if (not value == None and i > 1 and i < 7):
                ws.cell(row=startrow + cnt, column=i -
                        1).font = Font(color=color)
        cnt += 1

    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)
    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    
    return resp


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report2930(request):
    datastr = request.body
    data = json.loads(datastr)
    id = data['id']
    tip_rep = data['tip_rep']
    
    
    tipdoc = data['tipdoc']
    ids = [0, 0]
    if tipdoc == "svod":
        izmdocs = reg_svod_exp.objects.filter(_svod_exp_id = id)
        for itm in izmdocs:
            ids.append(itm._izm_exp_id)
    else:
        ids.append(id)

    # Получаем текущий каталог скрипта
    relative_path = os.path.join(
        current_directory, "report_template", "report_2930.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""WITH zapros as (SELECT * FROM public.docs_izm_exp_{tip_rep}
                            WHERE _izm_exp_id in {tuple(ids)}),
                    org as (SELECT * FROM public.dirs_organization
                            WHERE id in (select _organization_id from zapros)),
                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from zapros)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                            WHERE id in (select _spec_id from zapros)),
                    res as  (select 
                                    fg.code as fg_code, fg.name_rus as fg_name, 
                                    abp.code as abp_code, abp.name_rus as abp_name, 
                                    right(pr.code, 3) as pr_code, pr.name_rus as pr_name,
                                    right(ppr.code, 3) as ppr_code, ppr.name_rus as ppr_name,
                                    spec.code as spec_code, spec.name_rus as spec_name,
                                    sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12 as god,
                                    sm1,sm2,sm3,sm4,sm5,sm6,sm7,sm8,sm9,sm10,sm11,sm12
                                from zapros

                                left join fkr on
                                    zapros._fkr_id = fkr.id
                                left join fg on
                                    fkr._funcgroup = fg.id	
                                left join abp on
                                    fkr._abp = abp.id
                                left join pr on
                                    fkr._program_id = pr.id
                                left join ppr on
                                    fkr._podprogram_id = ppr.id
                                left join spec on
                                    zapros._spec_id = spec.id)
                                    
                SELECT  fg_code, abp_code, pr_code, ppr_code, spec_code, 
                        '' as name,
                        sum(god), sum(sm1), sum(sm2), sum(sm3), sum(sm4), sum(sm5), sum(sm6), sum(sm7), sum(sm8), sum(sm9), sum(sm10), sum(sm11), sum(sm12),
                        max(fg_name), max(abp_name),max(pr_name), max(ppr_name), max(spec_name)
                FROM res 
                GROUP BY ROLLUP (fg_code, abp_code, pr_code, ppr_code, spec_code)
                order by fg_code NULLS FIRST, abp_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    startrow = 14
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            continue

        ws.insert_rows(14 + cnt)
        for i, value in enumerate(row, 1):

            if i > 19:
                continue

            if i == 6:  # если это наименование
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            # cell = ws.cell(row=14 + cnt, column=i, value=value)
            # cell.border  = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 6):
                ws.cell(row=14 + cnt, column=i-1).font = Font(color=color)
        cnt += 1

    ws.print_title_rows = "17:17"

    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)
    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    
    return resp


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report3335(request):
    datastr = request.body
    data = json.loads(datastr)
    id = data['id']
    tip_rep = data['tip_rep']


    tipdoc = data['tipdoc']
    ids = [0, 0]
    if tipdoc == "svod":
        izmdocs = reg_svod_exp.objects.filter(_svod_exp_id = id)
        for itm in izmdocs:
            ids.append(itm._izm_exp_id)
    else:
        ids.append(id)


    # Получаем текущий каталог скрипта
    relative_path = os.path.join(current_directory, "report_template", "report_3335.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""WITH zapros as (SELECT * FROM public.docs_izm_exp_{tip_rep}
                            WHERE _izm_exp_id in {tuple(ids)}),
                    org as (SELECT * FROM public.dirs_organization
                            WHERE id in (select _organization_id from zapros)),
                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from zapros)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                            WHERE id in (select _spec_id from zapros)),
                    res as  (select 
                                    fg.code as fg_code, fg.name_rus as fg_name, 
                                    abp.code as abp_code, abp.name_rus as abp_name,
                                    org.codeorg as org_code,  org.name_rus as org_name,
                                    right(pr.code, 3) as pr_code, pr.name_rus as pr_name,
                                    right(ppr.code, 3) as ppr_code, ppr.name_rus as ppr_name,
                                    spec.code as spec_code, spec.name_rus as spec_name,
                                    sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12 as god,
                                    sm1,sm2,sm3,sm4,sm5,sm6,sm7,sm8,sm9,sm10,sm11,sm12
                                from zapros
                                left join org on
                                    zapros._organization_id = org.id
                                left join fkr on
                                    zapros._fkr_id = fkr.id
                                left join fg on
                                    fkr._funcgroup = fg.id	
                                left join abp on
                                    fkr._abp = abp.id
                                left join pr on
                                    fkr._program_id = pr.id
                                left join ppr on
                                    fkr._podprogram_id = ppr.id
                                left join spec on
                                    zapros._spec_id = spec.id)
                                    
                SELECT  '' as name, abp_code, org_code, pr_code, ppr_code, spec_code, 
                        sum(god), sum(sm1), sum(sm2), sum(sm3), sum(sm4), sum(sm5), sum(sm6), sum(sm7), sum(sm8), sum(sm9), sum(sm10), sum(sm11), sum(sm12),
                        max(abp_name), max(org_name), max(pr_name), max(ppr_name), max(spec_name)
                FROM res 
                GROUP BY ROLLUP (abp_code, org_code, pr_code, ppr_code, spec_code)
                order by abp_code NULLS FIRST, org_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    startrow = 21
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[1] == None):
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 19:
                continue

            if i == 1:  # если это наименование
                if (row[2] == None and row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
                # cell.alignment = cell.alignment.copy(wrap_text=True)
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 2 and i < 6):
                ws.cell(row=startrow + cnt, column=i -
                        1).font = Font(color=color)
        cnt += 1

    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)
    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    
    return resp


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_25(request):
    datastr = request.body
    data = json.loads(datastr)
    _izm_inc_id = data['_izm_inc_id']
    lang = data['lang']
    doc_obj = izm_inc.objects.get(id=_izm_inc_id)

    if (_izm_inc_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите бюджет"}), content_type="application/json", status=400)

    # Получаем текущий каталог скрипт
    relative_path = os.path.join(
        current_directory, "report_template", "report_25_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    if lang == 'kaz':
        ws._cells[15,8].value = "№" + doc_obj.nom.split('-')[0] + " АНЫКТАМА"
    else:
        ws._cells[14,8].value = "СПРАВКА №" + doc_obj.nom.split('-')[0]
    

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with reg as (SELECT * FROM docs_reg_inc where _izm_inc_id = {_izm_inc_id}),
                        budjet as (select * from dirs_budjet
                                    where id in (select _budjet_id from reg)),
                        classif as (select * from dirs_classification_income
                                    where id in (select _classification_id from reg)),
                        cat as (select * from dirs_category_income
                                    where id in (select _category_id from classif)),
                        clas as (select * from dirs_class_income
                                    where id in (select _classs_id from classif)),
                        podcl as (select * from dirs_podclass_income
                                    where id in (select _podclass_id from classif)),
                        spec as (select * from dirs_spec_income
                                    where id in (select _spec_id from classif))


                    select  '' as name,
                            cat.code as cat, right(clas.code, 2) as classs, right(podcl.code, 1) as podclass, right(spec.code,2) as spec,
                            sum(reg.god) as god, 
                            sum(reg.sm1) as sm1, 
                            sum(reg.sm2) as sm2, 
                            sum(reg.sm3) as sm3, 
                            sum(reg.sm4) as sm4, 
                            sum(reg.sm5) as sm5, 
                            sum(reg.sm6) as sm6, 
                            sum(reg.sm7) as sm7, 
                            sum(reg.sm8) as sm8, 
                            sum(reg.sm9) as sm9, 
                            sum(reg.sm10) as sm10, 
                            sum(reg.sm11) as sm11, 
                            sum(reg.sm12) as sm12,
                            max(cat.name_rus) as cat_name, max(clas.name_rus) as classs_name, max(podcl.name_rus) as podclass_name, max(spec.name_rus) as spec_name
                    from reg
                    left join budjet on 
                        reg._budjet_id = budjet.id
                    left join classif on 
                        reg._classification_id = classif.id
                    left join cat on 
                        classif._category_id = cat.id
                    left join clas on 
                        classif._classs_id = clas.id
                    left join podcl on 
                        classif._podclass_id = podcl.id
                    left join spec on 
                        classif._spec_id = spec.id
                    group by rollup(budjet.name_rus, cat.code, clas.code, podcl.code, spec.code)
                    order by budjet.name_rus nulls first, cat.code nulls first, clas.code nulls first, podcl.code nulls first, spec.code nulls first
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 20
    cnt = 0
    for row in result:
        color = "FFFFFF"
        # Пропуск общего итога
        if (row[1] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):
            # i начинается от 1 по ширине колонок (количество)
            if i > 18:
                continue

            if i == 1:  # если это наименование
                if (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[18])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 2 and i < 6):
                ws.cell(row=startrow + cnt, column=i -
                        1).font = Font(color=color)
        cnt += 1

    for i, value in enumerate(total_row, 1):
        # i начинается от 1 по ширине колонок (количество)
        if i > 18 or value == None:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    
    return resp


@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report3739(request):
    datastr = request.body
    data = json.loads(datastr)
    id = data['id']
    tip_rep = data['tip_rep']
    

    # Получаем текущий каталог скрипта
    relative_path = os.path.join(current_directory, "report_template", "report_3739_rus.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""WITH izm_svod as (select _izm_exp_id from docs_reg_svod_exp where _svod_exp_id = {id}),
					zapros as (SELECT * FROM public.docs_izm_exp_pay
                            WHERE _izm_exp_id in (select * from izm_svod)),
                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from zapros)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                            WHERE id in (select _funcgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                            WHERE id in (select _program_id from fkr)),
                    res as  (select 
                                    fg.code as fg_code, fg.name_rus as fg_name, 
                                    abp.code as abp_code, abp.name_rus as abp_name,
                                    right(pr.code, 3) as pr_code, pr.name_rus as pr_name,
                                    sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12 as god,
                                    sm1,sm2,sm3,sm4,sm5,sm6,sm7,sm8,sm9,sm10,sm11,sm12
                                from zapros
                                left join fkr on
                                    zapros._fkr_id = fkr.id
                                left join fg on
                                    fkr._funcgroup = fg.id	
                                left join abp on
                                    fkr._abp = abp.id
                                left join pr on
                                    fkr._program_id = pr.id)
                                    
                SELECT  '' as name, fg_code, abp_code, pr_code, 
                        sum(god), sum(sm1), sum(sm2), sum(sm3), sum(sm4), sum(sm5), sum(sm6), sum(sm7), sum(sm8), sum(sm9), sum(sm10), sum(sm11), sum(sm12),
                        max(fg_name), max(abp_name), max(pr_name)
                FROM res 
                GROUP BY ROLLUP (fg_code, abp_code, pr_code)
                order by fg_code NULLS FIRST, abp_code NULLS FIRST, pr_code NULLS FIRST"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    startrow = 23
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[1] == None):
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 17:
                continue

            if i == 1:  # если это наименование
                if (row[2] == None and row[3] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[17])
                elif (row[3] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[18])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            
            cell.border = border
            cell.alignment = Alignment(wrap_text=True)

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 2 and i < 5):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)
    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    
    return resp










# ********************************************
# *******ОТЧЕТЫ ПОЛУЧАЕМЫЕ С РЕГИСТРОВ********
# ********************************************
@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report1_4(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    tip_rep = data['tip_rep']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)


    if not (tip_rep == "pay" or tip_rep == "obl"):
        return HttpResponse(json.dumps({"status": "Неверный тип отчета"}), content_type="application/json", status=400)

    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)
    org_obj = organization.objects.get(id = org_id)

    


    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_1_4_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    
    ws._cells[10,16].value = str(datetime.date.today())
    
    if lang == 'kaz':
        ws._cells[12,6].value = org_obj._budjet.name_kaz
        ws._cells[13,6].value = org_obj._budjet._vid_budjet.name_kaz
        ws._cells[14,6].value = str(datetime.date.today())
        ws._cells[16,6].value = org_obj._abp.name_kaz
        ws._cells[17,6].value = org_obj.name_kaz
        if tip_rep == "pay":
            ws._cells[3,16].value = "1 косымша"
            ws._cells[9,9].value = "Төлемдер бойынша мемлекеттік мекемені қаржыландырудың жеке жобасы"
        else:
            ws._cells[3,16].value = "4 косымша"
            ws._cells[9,9].value = "Мiндеттеме бойынша мемлекеттік мекемені қаржыландырудың жеке жобасы"
    else:
        ws._cells[12,6].value = org_obj._budjet.name_rus
        ws._cells[13,6].value = org_obj._budjet._vid_budjet.name_rus
        ws._cells[14,6].value = str(datetime.date.today())
        ws._cells[16,6].value = org_obj._abp.name_rus
        ws._cells[17,6].value = org_obj.name_rus
        if tip_rep == "pay":
            ws._cells[1,16].value = "Приложение 1"
            ws._cells[9,9].value = "Проект индивидуального плана финансирования государственного учреждения по платежам"
        else:
            ws._cells[1,16].value = "Приложение 4"
            ws._cells[9,9].value = "Проект индивидуального плана финансирования государственного учреждения по обязательствам"
        


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with union_utv_izm as (SELECT _fkr_id, _spec_id, _organization_id, sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                            FROM public.docs_reg_exp_{tip_rep}
                            WHERE _organization_id = {org_id} and _date>='{date_start}' and _date<='{date_end}'
                            GROUP BY _fkr_id, _spec_id,_organization_id
                            HAVING not (sum(sm1)+sum(sm2)+sum(sm3)+sum(sm4)+sum(sm5)+sum(sm6)+sum(sm7)+sum(sm8)+sum(sm9)+sum(sm10)+sum(sm11)+sum(sm12))=0),
                    org as (SELECT * FROM public.dirs_organization
                                            WHERE id in (select _organization_id from union_utv_izm)),
                    fkr as (SELECT * FROM public.dirs_fkr
                                            WHERE id in (select _fkr_id from union_utv_izm)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                                            WHERE id in (select _spec_id from union_utv_izm))

                    SELECT  
                        abp.code as abp_code, 
                        org.codeorg as org_code,  
                        right(pr.code, 3) as pr_code, 
                        right(ppr.code, 3) as ppr_code, 
                        spec.code as spec_code, 
                        '' as name,
                        sum(sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12) as god,
                        COALESCE(sum(sm1),0) as sm1, 
                        COALESCE(sum(sm2),0) as sm2,
                        COALESCE(sum(sm3),0) as sm3, 
                        COALESCE(sum(sm4),0) as sm4, 
                        COALESCE(sum(sm5),0) as sm5, 
                        COALESCE(sum(sm6),0) as sm6, 
                        COALESCE(sum(sm7),0) as sm7, 
                        COALESCE(sum(sm8),0) as sm8, 
                        COALESCE(sum(sm9),0) as sm9, 
                        COALESCE(sum(sm10),0) as sm10, 
                        COALESCE(sum(sm11),0) as sm11, 
                        COALESCE(sum(sm12),0) as sm12,
                        max(abp.name_rus) as abp_name,
                        max(org.name_rus) as org_name,
                        max(pr.name_rus) as pr_name,
                        max(ppr.name_rus) as ppr_name,
                        max(spec.name_rus) as spec_name
                    FROM union_utv_izm as fulldata
                    left join org on
                        fulldata._organization_id = org.id
                    left join fkr on
                        fulldata._fkr_id = fkr.id
                    left join abp on
                        fkr._abp = abp.id
                    left join pr on
                        fkr._program_id = pr.id
                    left join ppr on
                        fkr._podprogram_id = ppr.id
                    left join spec on
                        fulldata._spec_id = spec.id
                    GROUP BY ROLLUP (abp_code, org_code, pr_code, ppr_code, spec_code)
                    order by abp_code NULLS FIRST, org_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 24
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 19:
                continue

            if i == 6:  # если это наименование
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 6):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value in enumerate(total_row, 1):
        if i > 19:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)
   
    # xlsx_temp = current_directory + "/temp_files/" + request.user.username + "_2_5.xlsx"
    # pdf_temp = current_directory + "/temp_files/" + request.user.username + "_2_5.pdf"
    # wb.save(xlsx_temp)
    # workbook = Workbook(xlsx_temp)
    # workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report2_5(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    tip_rep = data['tip_rep']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)


    if not (tip_rep == "pay" or tip_rep == "obl"):
        return HttpResponse(json.dumps({"status": "Неверный тип отчета"}), content_type="application/json", status=400)

    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)
    org_obj = organization.objects.get(id = org_id)

    


    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_2_5_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    
    ws._cells[10,16].value = str(datetime.date.today())
    
    if lang == 'kaz':
        ws._cells[12,6].value = org_obj._budjet.name_kaz
        ws._cells[13,6].value = org_obj._budjet._vid_budjet.name_kaz
        ws._cells[14,6].value = str(datetime.date.today())
        ws._cells[16,6].value = org_obj._abp.name_kaz
        ws._cells[17,6].value = org_obj.name_kaz
        if tip_rep == "pay":
            ws._cells[3,16].value = "2 косымша"
            ws._cells[9,9].value = "Төлемдер бойынша мемлекеттік мекемені қаржыландырудың жеке жоспары"
        else:
            ws._cells[3,16].value = "5 косымша"
            ws._cells[9,9].value = "Мiндеттеме бойынша мемлекеттік мекемені қаржыландырудың жеке жоспары"
    else:
        ws._cells[12,6].value = org_obj._budjet.name_rus
        ws._cells[13,6].value = org_obj._budjet._vid_budjet.name_rus
        ws._cells[14,6].value = str(datetime.date.today())
        ws._cells[16,6].value = org_obj._abp.name_rus
        ws._cells[17,6].value = org_obj.name_rus
        if tip_rep == "pay":
            ws._cells[1,16].value = "Приложение 2"
            ws._cells[9,9].value = "Индивидуальный план финансирования государственного учреждения по платежам"
        else:
            ws._cells[1,16].value = "Приложение 5"
            ws._cells[9,9].value = "Индивидуальный план финансирования государственного учреждения по обязательствам"
        


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with union_utv_izm as (SELECT _fkr_id, _spec_id, _organization_id, sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                            FROM public.docs_reg_exp_{tip_rep}
                            WHERE _organization_id = {org_id} and _date>='{date_start}' and _date<='{date_end}'
                            GROUP BY _fkr_id, _spec_id,_organization_id
                            HAVING not (sum(sm1)+sum(sm2)+sum(sm3)+sum(sm4)+sum(sm5)+sum(sm6)+sum(sm7)+sum(sm8)+sum(sm9)+sum(sm10)+sum(sm11)+sum(sm12))=0),
                    org as (SELECT * FROM public.dirs_organization
                                            WHERE id in (select _organization_id from union_utv_izm)),
                    fkr as (SELECT * FROM public.dirs_fkr
                                            WHERE id in (select _fkr_id from union_utv_izm)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                                            WHERE id in (select _spec_id from union_utv_izm))

                    SELECT  
                        abp.code as abp_code, 
                        org.codeorg as org_code,  
                        right(pr.code, 3) as pr_code, 
                        right(ppr.code, 3) as ppr_code, 
                        spec.code as spec_code, 
                        '' as name,
                        sum(sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12) as god,
                        COALESCE(sum(sm1),0) as sm1, 
                        COALESCE(sum(sm2),0) as sm2,
                        COALESCE(sum(sm3),0) as sm3, 
                        COALESCE(sum(sm4),0) as sm4, 
                        COALESCE(sum(sm5),0) as sm5, 
                        COALESCE(sum(sm6),0) as sm6, 
                        COALESCE(sum(sm7),0) as sm7, 
                        COALESCE(sum(sm8),0) as sm8, 
                        COALESCE(sum(sm9),0) as sm9, 
                        COALESCE(sum(sm10),0) as sm10, 
                        COALESCE(sum(sm11),0) as sm11, 
                        COALESCE(sum(sm12),0) as sm12,
                        max(abp.name_rus) as abp_name,
                        max(org.name_rus) as org_name,
                        max(pr.name_rus) as pr_name,
                        max(ppr.name_rus) as ppr_name,
                        max(spec.name_rus) as spec_name
                    FROM union_utv_izm as fulldata
                    left join org on
                        fulldata._organization_id = org.id
                    left join fkr on
                        fulldata._fkr_id = fkr.id
                    left join abp on
                        fkr._abp = abp.id
                    left join pr on
                        fkr._program_id = pr.id
                    left join ppr on
                        fkr._podprogram_id = ppr.id
                    left join spec on
                        fulldata._spec_id = spec.id
                    GROUP BY ROLLUP (abp_code, org_code, pr_code, ppr_code, spec_code)
                    order by abp_code NULLS FIRST, org_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 24
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 19:
                continue

            if i == 6:  # если это наименование
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 6):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value in enumerate(total_row, 1):
        if i > 19:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)


    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_isp_420(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    # tip_rep = data['tip_rep']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)

    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_isp_4_20_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

    # ws._cells[10,16].value = str(datetime.date.today())
    

    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')


    res_mas = [org_id]
    res = child_orgs([0, org_id])
    for i in range(1, 5, 1):
        massiv = [0]
        for itm in res:
            massiv.append(itm[0])
            res_mas.append(itm[0])
        if len(massiv)<2:
            break
        res = child_orgs(massiv)
    



    query = f"""with regdata as (SELECT _fkr_id, _spec_id, _organization_id, sum(god) as god_pl, 
                                    sum(sm1+sm2+sm3+sm4+sm5+sm6+sm7+sm8+sm9+sm10+sm11+sm12) as to_month_pl, sum(sm12) as cur_month_pl
                                FROM docs_reg_exp_pay
                                WHERE _organization_id in {tuple(res_mas)} and _date>='{date_start}' and _date<='{date_end}'
                                GROUP BY _fkr_id, _spec_id,_organization_id
                                HAVING not sum(god)=0),
				 
                    imp420 as (SELECT * FROM public.docs_import420 
                                WHERE _organization_id in {tuple(res_mas)} and not deleted and _date>='{date_start}' and _date<='{date_end}'),

                    date_by_org as (SELECT _organization_id, max(_date) as _date 
                                    FROM imp420 group by _organization_id),

                    actual420 as (SELECT * FROM imp420 
                                    WHERE (_organization_id, _date) in (select _organization_id, _date FROM date_by_org)),
                                    
                    tbl420 as (SELECT _organization_id, _fkr_id, _spec_id, sum(sm3/1000) as to_month420, sum(sm9/1000) as ost_obl_420, sum(sm10/1000) as ost_pay_420 
                                FROM public.docs_import420_tbl1 
                                WHERE _import420_id in (select id FROM actual420)
                                GROUP BY _organization_id, _fkr_id, _spec_id),
				
                    org as (SELECT id, codeorg, name_rus, _region_id FROM public.dirs_organization
                            WHERE id in (select _organization_id from regdata)),

                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from regdata)),

                    abp as (SELECT * FROM public.dirs_abp
                            WHERE id in (select _abp from fkr)),

                    pr  as (SELECT * FROM public.dirs_program
                            WHERE id in (select _program_id from fkr)),

                    ppr as (SELECT * FROM public.dirs_podprogram
                            WHERE id in (select _podprogram_id from fkr)),

                    spec as (SELECT * FROM public.dirs_spec_exp
                            WHERE id in (select _spec_id from regdata))

                    SELECT  
						org._region_id as region,
						org.codeorg as org_code,
                        abp.code as abp_code,      
                        right(pr.code, 3) as pr_code, 
                        right(ppr.code, 3) as ppr_code, 
                        spec.code as spec_code, 
                        '' as name,
                        sum(god_pl) as god_pl,
                        COALESCE(sum(to_month_pl),0) as to_month_pl, 
                        COALESCE(sum(cur_month_pl),0) as cur_month_pl,
						COALESCE(sum(to_month420)) as to_month420,
						(COALESCE(sum(god_pl),0)-COALESCE(sum(to_month420),0)) as kassa,
                        COALESCE(sum(ost_obl_420),0) as ost_obl_420, 
                        COALESCE(sum(ost_pay_420),0) as ost_pay_420,
                        ceil(COALESCE(COALESCE(sum(to_month420)) * 100/COALESCE(sum(to_month_pl),0),0)) as perc,
                        max(org.name_rus) as org_name,
                        max(abp.name_rus) as abp_name,
                        max(pr.name_rus) as pr_name,
                        max(ppr.name_rus) as ppr_name,
                        max(spec.name_rus) as spec_name
                    FROM regdata as fulldata
					LEFT JOIN tbl420 on
						fulldata._organization_id = tbl420._organization_id and 
						fulldata._fkr_id = tbl420._fkr_id and
						fulldata._spec_id = tbl420._spec_id
                    LEFT JOIN org on
                        fulldata._organization_id = org.id
                    LEFT JOIN fkr on
                        fulldata._fkr_id = fkr.id
                    LEFT JOIN abp on
                        fkr._abp = abp.id
                    LEFT JOIN pr on
                        fkr._program_id = pr.id
                    LEFT JOIN ppr on
                        fkr._podprogram_id = ppr.id
                    LEFT JOIN spec on
                        fulldata._spec_id = spec.id
                    GROUP BY ROLLUP (region, org_code, abp_code, pr_code, ppr_code, spec_code)
                    ORDER BY region NULLS FIRST,  org_code NULLS FIRST, abp_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST, spec_code NULLS FIRST          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 7
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 15:
                continue

            if i == 7:  # если это наименование
                ф = 1
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[0])
                elif (row[2] == None and row[3] == None and row[4] == None and row[5] == None):
                    # cell = ws.cell(row=startrow + cnt, column=i, value="")
                    cell = ws.cell(row=startrow + cnt, column=2, value=row[15])
                elif (row[3] == None and row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[16])
                elif (row[4] == None and row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[17])
                elif (row[5] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[18])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 7):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value in enumerate(total_row, 1):
        if i<=7 or i>15:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    ws.print_area = ws.calculate_dimension()
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)


    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp




@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_isp_219(request):
    datastr = request.body
    data = json.loads(datastr)
    budjet_id = data['_budjet_id']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)


    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_isp_2_19_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    curbdj = budjet.objects.get(id = budjet_id)
    ids = curbdj.get_hierarchy_ids(include_self=True)
    ids.append(0)


    query = f"""with    reg as (SELECT _budjet_id,
                                _classification_id,
                                sum(reg.god) as god_pl, 
                                sum(reg.sm1+reg.sm2+reg.sm3+reg.sm4+reg.sm5+reg.sm6+reg.sm7+reg.sm8+reg.sm9+reg.sm10+reg.sm11+reg.sm12) as to_month_pl
                            FROM docs_reg_inc as reg
                            WHERE _budjet_id in {tuple(ids)} and _date>='{date_start}' and _date<='{date_end}'   
                            GROUP BY _budjet_id, _classification_id
                            HAVING not sum(reg.god)=0
                            ),

	                    imp219 as (SELECT * FROM public.docs_import219 
                                    WHERE not deleted and _date>='{date_start}' and _date<='{date_end}'
                                        and _budjet_id in {tuple(ids)}),

                        date_by_org as (SELECT _budjet_id, max(_date) as _date 
                                        FROM imp219 
                                        GROUP BY _budjet_id),

                        actual219 as (SELECT * FROM imp219  
                                        WHERE (_budjet_id, _date) in (select _budjet_id, _date FROM date_by_org)),						
	 
                        tbl219 as (SELECT _budjet_id, _classification_id,
                                        sum(reg.sm6/1000) as god_219,
                                        sum(reg.sm9/1000) as ost_obl
                                    FROM docs_import219_tbl1 as reg 
							        WHERE _import219_id in (select id FROM actual219)
                                    GROUP BY _budjet_id, _classification_id),
							
                        unionall as (SELECT _budjet_id, _classification_id, god_pl, to_month_pl, 0 as god_219, 0 as ost_obl FROM reg
                                        UNION ALL
                                    SELECT _budjet_id, _classification_id, 0, 0, god_219, ost_obl FROM tbl219),

                        groupunionall as (SELECT _budjet_id, _classification_id, 
                                            sum(god_pl) as god_pl, 
                                            sum(to_month_pl) as to_month_pl, 
                                            sum(god_219) as god_219,
                                            (sum(god_pl) - sum(god_219)) as ost_god,
                                            sum(ost_obl) as ost_obl 
                                            FROM unionall
                                            GROUP BY _budjet_id, _classification_id),
					 
                        budjet as (select * from dirs_budjet
                                    where id in (select _budjet_id from groupunionall)),

                        classif as (select * from dirs_classification_income
                                    where id in (select _classification_id from groupunionall)),

                        cat as (select * from dirs_category_income
                                where id in (select _category_id from classif)),

                        clas as (select * from dirs_class_income
                                where id in (select _classs_id from classif)),

                        podcl as (select * from dirs_podclass_income
                                where id in (select _podclass_id from classif)),

                        spec as (select * from dirs_spec_income
                                where id in (select _spec_id from classif))
									
	            SELECT  budjet.name_rus as budjet,
                        cat.code as cat, 
                        right(clas.code, 2) as classs, 
                        right(podcl.code, 1) as podclass, 
                        right(spec.code,2) as spec, 
                        '' as name,
                        sum(groupunionall.god_pl) as god_pl, 
                        sum(groupunionall.to_month_pl) as to_month_pl, 
                        sum(groupunionall.god_219) as god_219, 
                        sum(groupunionall.ost_god) as ost_god,
						sum(groupunionall.ost_obl) as ost_obl,
                        case 
                            when sum(groupunionall.god_pl)=0
                                then 0
                            else 
                                sum(groupunionall.god_219) * 100/sum(groupunionall.god_pl) 
                        end as proc,
						max(cat.name_rus) as cat_name, 
                        max(clas.name_rus) as classs_name, 
                        max(podcl.name_rus) as podclass_name, 
                        max(spec.name_rus) as spec_name
                FROM groupunionall
                LEFT JOIN budjet on 
                        groupunionall._budjet_id = budjet.id
                LEFT JOIN classif on 
                        groupunionall._classification_id = classif.id
                LEFT JOIN cat on 
                        classif._category_id = cat.id
                LEFT JOIN clas on 
                        classif._classs_id = clas.id
                LEFT JOIN podcl on 
                        classif._podclass_id = podcl.id
                LEFT JOIN spec on 
                        classif._spec_id = spec.id
                GROUP BY rollup(budjet.name_rus, cat.code, clas.code, podcl.code, spec.code)
                ORDER BY budjet.name_rus nulls first, cat.code nulls first, clas.code nulls first, podcl.code nulls first, spec.code nulls first          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 8
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 12:
                continue

            if i == 6:  # если это наименование
                ф = 1
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None ):
                    cell = ws.cell(row=startrow + cnt, column=i, value='')
                elif (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[12])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[13])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[14])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[15])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 5):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value_ in enumerate(total_row, 1):
        if i<=6 or i > 12:
            continue
        ws.cell(row=startrow + cnt, column=i, value=value_)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    ws.print_area = ws.calculate_dimension()
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)


    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp




@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_14(request):
    datastr = request.body
    data = json.loads(datastr)
    _budjet_id = data['_budjet_id']
    _date = data['_date']
    lang = data['lang']

    date = datetime.datetime.strptime(_date, '%d.%m.%Y %H:%M:%S')
    date_end = date.replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)
    bjt_obj = budjet.objects.get(id=_budjet_id)
    ids_bjt = bjt_obj.get_hierarchy_ids(include_self=True)
    ids_bjt.append(0)  # пустой ИД в конце, чтобы избавиться от лишнего ","
    kortej = tuple(ids_bjt)

    if (_budjet_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите бюджет"}), content_type="application/json", status=400)

    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_14_" + lang + ".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active


    if lang == 'kaz':
        ws._cells[13,6].value = bjt_obj._vid_budjet.name_kaz
        ws._cells[14,6].value = date.year
        ws._cells[15,6].value = date.date().strftime('%d.%m.%Y')
    else:
        ws._cells[13,6].value = bjt_obj._vid_budjet.name_rus
        ws._cells[14,6].value = date.year
        ws._cells[15,6].value = date.date().strftime('%d.%m.%Y')


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with reg as (SELECT _budjet_id,
                                _classification_id,
                                sum(reg.god) as god, 
                                sum(reg.sm1) as sm1, 
                                sum(reg.sm2) as sm2, 
                                sum(reg.sm3) as sm3, 
                                sum(reg.sm4) as sm4, 
                                sum(reg.sm5) as sm5, 
                                sum(reg.sm6) as sm6, 
                                sum(reg.sm7) as sm7, 
                                sum(reg.sm8) as sm8, 
                                sum(reg.sm9) as sm9, 
                                sum(reg.sm10) as sm10, 
                                sum(reg.sm11) as sm11, 
                                sum(reg.sm12) as sm12
                            FROM docs_reg_inc as reg
                            WHERE _budjet_id in {kortej}
                                and _date>='{date_start}' and _date<='{date_end}'
                            GROUP BY _budjet_id, _classification_id
                            HAVING not sum(reg.god)=0
                            ),
                        budjet as (select * from dirs_budjet
                                    where id in (select _budjet_id from reg)),
                        classif as (select * from dirs_classification_income
                                    where id in (select _classification_id from reg)),
                        cat as (select * from dirs_category_income
                                    where id in (select _category_id from classif)),
                        clas as (select * from dirs_class_income
                                    where id in (select _classs_id from classif)),
                        podcl as (select * from dirs_podclass_income
                                    where id in (select _podclass_id from classif)),
                        spec as (select * from dirs_spec_income
                                    where id in (select _spec_id from classif))


                    select  budjet.name_rus as budjet,
                            cat.code as cat, right(clas.code, 2) as classs, right(podcl.code, 1) as podclass, right(spec.code,2) as spec, 
                            '' as name,
                            sum(reg.god) as god, 
                            sum(reg.sm1) as sm1, 
                            sum(reg.sm2) as sm2, 
                            sum(reg.sm3) as sm3, 
                            sum(reg.sm4) as sm4, 
                            sum(reg.sm5) as sm5, 
                            sum(reg.sm6) as sm6, 
                            sum(reg.sm7) as sm7, 
                            sum(reg.sm8) as sm8, 
                            sum(reg.sm9) as sm9, 
                            sum(reg.sm10) as sm10, 
                            sum(reg.sm11) as sm11, 
                            sum(reg.sm12) as sm12,
                            '' as bjt_name, max(cat.name_rus) as cat_name, max(clas.name_rus) as classs_name, max(podcl.name_rus) as podclass_name, max(spec.name_rus) as spec_name
                    from reg
                    left join budjet on 
                        reg._budjet_id = budjet.id
                    left join classif on 
                        reg._classification_id = classif.id
                    left join cat on 
                        classif._category_id = cat.id
                    left join clas on 
                        classif._classs_id = clas.id
                    left join podcl on 
                        classif._podclass_id = podcl.id
                    left join spec on 
                        classif._spec_id = spec.id
                    group by rollup(budjet.name_rus, cat.code, clas.code, podcl.code, spec.code)
                    order by budjet.name_rus nulls first, cat.code nulls first, clas.code nulls first, podcl.code nulls first, spec.code nulls first
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 22
    cnt = 0
    for row in result:
        color = "FFFFFF"
        # Пропуск общего итога
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):
            if i > 19:
                continue

            if i == 6:  # если это наименование
                if (row[1] == None and row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
                elif (row[2] == None and row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
                elif (row[3] == None and row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
                elif (row[4] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 6):
                ws.cell(row=startrow + cnt, column=i -
                        1).font = Font(color=color)
        cnt += 1

    # Общие итоги по всей таблице
    for i, value in enumerate(total_row, 1):
        if i > 19 or i<7:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)

    # xlsx_temp = current_directory + "/temp_files/" + \
    #     request.user.username + "_14.xlsx"
    # pdf_temp = current_directory + "/temp_files/" + request.user.username + "_14.pdf"
    # wb.save(xlsx_temp)
    # workbook = Workbook(xlsx_temp)
    # workbook.save(pdf_temp)

    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)

    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'
    
    os.close(id_xslx)
    os.close(id_pdf)
    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report8_10(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    tip_rep = data['tip_rep']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)

    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)
    
    mass_ids = []
    mass_ids.append(0)
    mass_ids.append(org_id)
    child_orgs = parent_organizations.objects.filter(_parent_id = org_id).values('_organization_id')
    for item in child_orgs:
        mass_ids.append(item['_organization_id'])
    org_obj = organization.objects.get(id=org_id)
        
      
    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_8_10_"+lang+".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

       
    if lang == 'kaz':
        ws._cells[14,6].value = org_obj._budjet._vid_budjet.name_kaz
        ws._cells[15,6].value = str(datetime.date.today())
        ws._cells[17,6].value = org_obj._abp.name_kaz

        if tip_rep == "pay":
            ws._cells[3,16].value = "8 косымша"
            ws._cells[9,9].value = "Төлемдер бойынша мемлекеттік мекемені қаржыландырудың жоспары"
        else:
            ws._cells[3,16].value = "10 косымша"
            ws._cells[9,9].value = "Мiндеттеме бойынша мемлекеттік мекемені қаржыландырудың жоспары"
    else:
        ws._cells[14,6].value = org_obj._budjet._vid_budjet.name_rus
        ws._cells[15,6].value = str(datetime.date.today())
        ws._cells[17,6].value = org_obj._abp.name_rus
        if tip_rep == "pay":
            ws._cells[1,16].value = "Приложение 8"
            ws._cells[9,9].value = "План финансирования государственного учреждения по платежам"
        else:
            ws._cells[1,16].value = "Приложение 10"
            ws._cells[9,9].value = "План финансирования государственного учреждения по обязательствам"
        


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')

    query = f"""with 
                    svod_docs as (SELECT _izm_exp_id FROM public.docs_reg_svod_exp where _organization_id = {org_id} and _date>='{date_start}' and _date<='{date_end}'),    
                    union_utv_izm as (SELECT _fkr_id, _spec_id, _organization_id, sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                            FROM public.docs_reg_exp_{tip_rep}
                            WHERE (_izm_exp_id in (select _izm_exp_id from svod_docs)) or (_izm_exp_id is null and _organization_id in {tuple(mass_ids)})
                            or (_organization_id = {org_id})
                            GROUP BY _fkr_id, _spec_id,_organization_id
                            HAVING not (sum(sm1)+sum(sm2)+sum(sm3)+sum(sm4)+sum(sm5)+sum(sm6)+sum(sm7)+sum(sm8)+sum(sm9)+sum(sm10)+sum(sm11)+sum(sm12))=0),
                    org as (SELECT * FROM public.dirs_organization
                                            WHERE id in (select _organization_id from union_utv_izm)),
                    fkr as (SELECT * FROM public.dirs_fkr
                                            WHERE id in (select _fkr_id from union_utv_izm)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                                            WHERE id in (select _spec_id from union_utv_izm))

                    SELECT  
                        abp.code as abp_code, 
                        right(pr.code, 3) as pr_code, 
                        right(ppr.code, 3) as ppr_code, 
                        '' as name,
                        sum(sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12) as god,
                        COALESCE(sum(sm1),0) as sm1, 
                        COALESCE(sum(sm2),0) as sm2,
                        COALESCE(sum(sm3),0) as sm3, 
                        COALESCE(sum(sm4),0) as sm4, 
                        COALESCE(sum(sm5),0) as sm5, 
                        COALESCE(sum(sm6),0) as sm6, 
                        COALESCE(sum(sm7),0) as sm7, 
                        COALESCE(sum(sm8),0) as sm8, 
                        COALESCE(sum(sm9),0) as sm9, 
                        COALESCE(sum(sm10),0) as sm10, 
                        COALESCE(sum(sm11),0) as sm11, 
                        COALESCE(sum(sm12),0) as sm12,
                        max(abp.name_rus) as abp_name,
                        max(pr.name_rus) as pr_name,
                        max(ppr.name_rus) as ppr_name
                    FROM union_utv_izm as fulldata
                    left join fkr on
                        fulldata._fkr_id = fkr.id
                    left join abp on
                        fkr._abp = abp.id
                    left join pr on
                        fkr._program_id = pr.id
                    left join ppr on
                        fkr._podprogram_id = ppr.id
                    GROUP BY ROLLUP (abp_code, pr_code, ppr_code)
                    order by abp_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 23
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 17:
                continue

            if i == 4:  # если это наименование
                if (row[1] == None and row[2] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[17])
                elif (row[2] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[18])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 4):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value in enumerate(total_row, 1):
        if i > 17 or i < 4:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)

    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report7_9(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    tip_rep = data['tip_rep']
    lang = data['lang']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)

    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)
    
    mass_ids = []
    mass_ids.append(0)
    mass_ids.append(org_id)
    child_orgs = parent_organizations.objects.filter(_parent_id = org_id).values('_organization_id')
    for item in child_orgs:
        mass_ids.append(item['_organization_id'])
    org_obj = organization.objects.get(id=org_id)
        
      
    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_7_9_"+lang+".xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

       
    if lang == 'kaz':
        ws._cells[14,6].value = org_obj._budjet._vid_budjet.name_kaz
        ws._cells[15,6].value = str(datetime.date.today())
        ws._cells[17,6].value = org_obj._abp.name_kaz

        if tip_rep == "pay":
            ws._cells[3,16].value = "7 косымша"
            ws._cells[9,9].value = "Төлемдер бойынша мемлекеттік мекемені қаржыландырудың жобасы"
        else:
            ws._cells[3,16].value = "9 косымша"
            ws._cells[9,9].value = "Мiндеттеме бойынша мемлекеттік мекемені қаржыландырудың жобасы"
    else:
        ws._cells[14,6].value = org_obj._budjet._vid_budjet.name_rus
        ws._cells[15,6].value = str(datetime.date.today())
        ws._cells[17,6].value = org_obj._abp.name_rus
        if tip_rep == "pay":
            ws._cells[1,16].value = "Приложение 7"
            ws._cells[9,9].value = "Проект плана финансирования государственного учреждения по платежам"
        else:
            ws._cells[1,16].value = "Приложение 9"
            ws._cells[9,9].value = "Проект плана финансирования государственного учреждения по обязательствам"
        


    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD', end_color='FFDDDDDD', fill_type='solid')

    query = f"""with 
                    svod_docs as (SELECT _izm_exp_id FROM public.docs_reg_svod_exp where _organization_id = {org_id} and _date>='{date_start}' and _date<='{date_end}'),    
                    union_utv_izm as (SELECT _fkr_id, _spec_id, _organization_id, sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12 
                            FROM public.docs_reg_exp_{tip_rep}
                            WHERE (_izm_exp_id in (select _izm_exp_id from svod_docs)) or (_izm_exp_id is null and _organization_id in {tuple(mass_ids)})
                            or (_organization_id = {org_id})
                            GROUP BY _fkr_id, _spec_id,_organization_id
                            HAVING not (sum(sm1)+sum(sm2)+sum(sm3)+sum(sm4)+sum(sm5)+sum(sm6)+sum(sm7)+sum(sm8)+sum(sm9)+sum(sm10)+sum(sm11)+sum(sm12))=0),
                    org as (SELECT * FROM public.dirs_organization
                                            WHERE id in (select _organization_id from union_utv_izm)),
                    fkr as (SELECT * FROM public.dirs_fkr
                                            WHERE id in (select _fkr_id from union_utv_izm)),
                    fg as (SELECT * FROM public.dirs_funcgroup
                                            WHERE id in (select _funcgroup from fkr)),
                    fpg as (SELECT * FROM public.dirs_funcpodgroup
                                            WHERE id in (select _funcpodgroup from fkr)),
                    abp as (SELECT * FROM public.dirs_abp
                                            WHERE id in (select _abp from fkr)),
                    pr  as (SELECT * FROM public.dirs_program
                                            WHERE id in (select _program_id from fkr)),
                    ppr as (SELECT * FROM public.dirs_podprogram
                                            WHERE id in (select _podprogram_id from fkr)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                                            WHERE id in (select _spec_id from union_utv_izm))

                    SELECT  
                        abp.code as abp_code, 
                        right(pr.code, 3) as pr_code, 
                        right(ppr.code, 3) as ppr_code, 
                        '' as name,
                        sum(sm1 + sm2 + sm3 + sm4 + sm5 + sm6 + sm7 + sm8 + sm9 + sm10 + sm11 + sm12) as god,
                        COALESCE(sum(sm1),0) as sm1, 
                        COALESCE(sum(sm2),0) as sm2,
                        COALESCE(sum(sm3),0) as sm3, 
                        COALESCE(sum(sm4),0) as sm4, 
                        COALESCE(sum(sm5),0) as sm5, 
                        COALESCE(sum(sm6),0) as sm6, 
                        COALESCE(sum(sm7),0) as sm7, 
                        COALESCE(sum(sm8),0) as sm8, 
                        COALESCE(sum(sm9),0) as sm9, 
                        COALESCE(sum(sm10),0) as sm10, 
                        COALESCE(sum(sm11),0) as sm11, 
                        COALESCE(sum(sm12),0) as sm12,
                        max(abp.name_rus) as abp_name,
                        max(pr.name_rus) as pr_name,
                        max(ppr.name_rus) as ppr_name
                    FROM union_utv_izm as fulldata
                    left join fkr on
                        fulldata._fkr_id = fkr.id
                    left join abp on
                        fkr._abp = abp.id
                    left join pr on
                        fkr._program_id = pr.id
                    left join ppr on
                        fkr._podprogram_id = ppr.id
                    GROUP BY ROLLUP (abp_code, pr_code, ppr_code)
                    order by abp_code NULLS FIRST, pr_code NULLS FIRST, ppr_code NULLS FIRST          
                """

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 23
    cnt = 0
    for row in result:
        color = "FFFFFF"
        if (row[0] == None):
            total_row = row
            continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            if i > 17:
                continue

            if i == 4:  # если это наименование
                if (row[1] == None and row[2] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[17])
                elif (row[2] == None):
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[18])
                else:
                    cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
            else:
                cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            if (not value == None and i > 1 and i < 4):
                ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    for i, value in enumerate(total_row, 1):
        if i > 17 or i < 4:
            continue
        cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)

    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_diff_pay_obl(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)


    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)
    # org_obj = organization.objects.get(id = org_id)

    


    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_prev_pay_and_obl.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

 
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with reg_pay as (SELECT _fkr_id, _spec_id, _organization_id, sum(god) as god, 
                                    sum(sm1) as sm1, 
                                    sum(sm2 + sm1) as sm2, 
                                    sum(sm3 + sm2 + sm1) as sm3, 
                                    sum(sm4 + sm3 + sm2 + sm1) as sm4, 
                                    sum(sm5 + sm4 + sm3 + sm2 + sm1) as sm5, 
                                    sum(sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm6, 
                                    sum(sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm7, 
                                    sum(sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm8, 
                                    sum(sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm9, 
                                    sum(sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm10, 
                                    sum(sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm11, 
                                    sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm12 
                                FROM public.docs_reg_exp_pay
                                WHERE _organization_id = {org_id}
                                GROUP BY _fkr_id, _spec_id,_organization_id),
                    reg_obl as (SELECT _fkr_id, _spec_id, _organization_id, sum(god) as god, 
                                    sum(sm1) as sm1, 
                                    sum(sm2 + sm1) as sm2, 
                                    sum(sm3 + sm2 + sm1) as sm3, 
                                    sum(sm4 + sm3 + sm2 + sm1) as sm4, 
                                    sum(sm5 + sm4 + sm3 + sm2 + sm1) as sm5, 
                                    sum(sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm6, 
                                    sum(sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm7, 
                                    sum(sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm8, 
                                    sum(sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm9, 
                                    sum(sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm10, 
                                    sum(sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm11, 
                                    sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as sm12 
                                FROM public.docs_reg_exp_obl
                                WHERE _organization_id = {org_id}
                                GROUP BY _fkr_id, _spec_id,_organization_id),
                    unionall as (select _fkr_id, _spec_id, _organization_id, sm12 as god, sm1, sm2, sm3, sm4, sm5, sm6, sm7, sm8, sm9, sm10, sm11, sm12 from reg_obl
                                union all
                                select _fkr_id, _spec_id, _organization_id, -sm12 as god, -sm1, -sm2, -sm3, -sm4, -sm5, -sm6, -sm7, -sm8, -sm9, -sm10, -sm11, -sm12 from reg_pay),
                    
                    fkr as (SELECT * FROM public.dirs_fkr
                            WHERE id in (select _fkr_id from unionall)),
                    spec as (SELECT * FROM public.dirs_spec_exp
                            WHERE id in (select _spec_id from unionall))
                                                    
                    
                    select 	max(SPLIT_PART(fkr.code, '/', 1)) as abp,
                                max(SPLIT_PART(fkr.code, '/', 2)) as pr,
                                max(SPLIT_PART(fkr.code, '/', 3)) as ppr,
                                max(spec.code) as spec,
                                max(spec.name_rus) as spec_name,
                                sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, 
                                sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12, sum(sm12) as god
                    from unionall
                    left join fkr on
                            unionall._fkr_id=fkr.id
                    left join spec on
                            unionall._spec_id=spec.id
                    group by _fkr_id, _spec_id, _organization_id
                    having not (sum(sm1)=0 and sum(sm2)=0 and sum(sm3)=0 and sum(sm4)=0 and sum(sm5)=0 and sum(sm6)=0 and 
                                sum(sm7)=0 and sum(sm8)=0 and sum(sm9)=0 and sum(sm10)=0 and sum(sm11)=0 and sum(sm12)=0)"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()

    total_row = None
    startrow = 6
    cnt = 0
    for row in result:
        color = "FFFFFF"
        # if (row[0] == None):
        #     total_row = row
        #     continue

        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):

            # if i > 19:
            #     continue

            # if i == 6:  # если это наименование
            #     if (row[1] == None and row[2] == None and row[3] == None and row[4] == None):
            #         cell = ws.cell(row=startrow + cnt, column=i, value=row[19])
            #     elif (row[2] == None and row[3] == None and row[4] == None):
            #         cell = ws.cell(row=startrow + cnt, column=i, value=row[20])
            #     elif (row[3] == None and row[4] == None):
            #         cell = ws.cell(row=startrow + cnt, column=i, value=row[21])
            #     elif (row[4] == None):
            #         cell = ws.cell(row=startrow + cnt, column=i, value=row[22])
            #     else:
            #         cell = ws.cell(row=startrow + cnt, column=i, value=row[23])
            # else:
                
            cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"

            # if (not value == None and i > 1 and i < 6):
            #     ws.cell(row=startrow + cnt, column=i - 1).font = Font(color=color)
        cnt += 1

    # Вывод общих  итогов таблицы
    # for i, value in enumerate(total_row, 1):
    #     if i > 19:
    #         continue
    #     cell = ws.cell(row=startrow + cnt, column=i, value=value)


    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)
   
    # xlsx_temp = current_directory + "/temp_files/" + request.user.username + "_2_5.xlsx"
    # pdf_temp = current_directory + "/temp_files/" + request.user.username + "_2_5.pdf"
    # wb.save(xlsx_temp)
    # workbook = Workbook(xlsx_temp)
    # workbook.save(pdf_temp)

    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp



@api_view(['POST', 'GET'])
@permission_classes([IsAuthenticated])
def report_diff_god_sum(request):
    datastr = request.body
    data = json.loads(datastr)
    org_id = data['_organization_id']
    _date = data['_date']
    date_end = datetime.datetime.strptime(_date, "%d.%m.%Y %H:%M:%S").replace(hour=23, minute=59, second=59)
    date_start = date_end.replace(month=1, day=1, hour=0, minute=0, second=0)


    if (org_id == 0):
        return HttpResponse(json.dumps({"status": "Выберите организацию"}), content_type="application/json", status=400)


    # Получаем текущий каталог скрипт
    relative_path = os.path.join(current_directory, "report_template", "report_diff_god_sum.xlsx")
    wb = load_workbook(relative_path)
    ws = wb.active

 
    border = Border(left=Side(style='thin', color='000000'),
                    right=Side(style='thin', color='000000'),
                    top=Side(style='thin', color='000000'),
                    bottom=Side(style='thin', color='000000'))

    light_gray_fill = PatternFill(start_color='FFDDDDDD',
                                  end_color='FFDDDDDD',
                                  fill_type='solid')

    query = f"""with reg_pay as (SELECT _fkr_id, _spec_id, _organization_id, 
				 	case
				 		when _izm_exp_id IS NULL
				 			then sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1)
						else
				 			0
					end as utv_pay,
				 	sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as utoch_pay
                  FROM public.docs_reg_exp_pay
                  WHERE 1 = 1
                  GROUP BY _izm_exp_id, _fkr_id, _spec_id,_organization_id),
	reg_obl as (SELECT _fkr_id, _spec_id, _organization_id, 
					case
				 		when _izm_exp_id is null
				 			then sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1)
						else
				 			0
					end as utv_obl,
				 	sum(sm12 + sm11 + sm10 + sm9 + sm8 + sm7 + sm6 + sm5 + sm4 + sm3 + sm2 + sm1) as utoch_obl 
                  FROM public.docs_reg_exp_obl
                  WHERE 1 = 1
                  GROUP BY _izm_exp_id, _fkr_id, _spec_id,_organization_id),
    unionall as (select _fkr_id, _spec_id, _organization_id, utv_pay, utoch_pay, 0 as utv_obl, 0 as utoch_obl from reg_pay
				 union all
				 select _fkr_id, _spec_id, _organization_id, 0, 0, utv_obl, utoch_obl from reg_obl),
	
	fkr as (SELECT * FROM public.dirs_fkr
             WHERE id in (select _fkr_id from unionall)),
	spec as (SELECT * FROM public.dirs_spec_exp
             WHERE id in (select _spec_id from unionall))
									
		select 	max(SPLIT_PART(fkr.code, '/', 1)) as abp,
				max(SPLIT_PART(fkr.code, '/', 2)) as pr,
				max(SPLIT_PART(fkr.code, '/', 3)) as ppr,
				max(spec.code) as spec,
				sum(utv_pay) as utv_pay,
				sum(utv_obl) as utv_obl,
				sum(utoch_pay) as utoch_pay,
				sum(utoch_obl) as utoch_obl
		from unionall
		left join fkr on
			unionall._fkr_id=fkr.id
		left join spec on
			unionall._spec_id=spec.id
		group by _fkr_id, _spec_id, _organization_id"""

    with connection.cursor() as cursor:
        cursor.execute(query)
        result = cursor.fetchall()


    startrow = 8
    cnt = 0
    for row in result:
        ws.insert_rows(startrow + cnt)
        for i, value in enumerate(row, 1):
            cell = ws.cell(row=startrow + cnt, column=i, value=value)
            cell.border = border

            if cnt % 2 == 0:
                cell.fill = light_gray_fill
                color = "FFDDDDDD"
        cnt += 1



    id_xslx, xlsx_temp = tempfile.mkstemp(suffix='.xlsx', text=False)
    id_pdf, pdf_temp = tempfile.mkstemp(suffix='.pdf', text=False)
    wb.save(xlsx_temp)
    workbook = Workbook(xlsx_temp)
    workbook.save(pdf_temp)
   
    pdf_file = open(pdf_temp, "rb")
    body = FileWrapper(pdf_file)
    resp = HttpResponse(body, content_type="application/pdf")
    resp['Content-Disposition'] = 'attachment; filename="table.pdf"'

    os.close(id_xslx)
    os.close(id_pdf)

    os.remove(xlsx_temp)
    os.remove(pdf_temp)
    return resp





