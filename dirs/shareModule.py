from datetime import datetime, timedelta
from django.db import connection, transaction
from .models import *
from openpyxl import load_workbook
import tabula
import os


def fkrreadxls123(path='fkrnewkaz.xlsx'):
    listnum = ('0', '1', '2', '3', '4', '5', '6', '7','8', '9')
    
    if os.path.exists(path):  
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        first_sheet = workbook[sheet_names[0]]


        fg_id = 0
        fg_code = ''
        fpg_id = 0
        fpg_code = ''
        abp_id = 0
        abp_code = ''
        pr_id = 0
        pr_code = ''
        pr_name = ''
        ppr_id = 0
        ppr_code = ''

        lang = ''
        if first_sheet._cells[1,1].value == 'Функциональная группа':
            lang = 'rus'

        if first_sheet._cells[1,1].value == 'Функционалдық топ':
            lang = 'kaz'

        if lang == '':
            return 'error excel'
       

        if lang == 'kaz':
            # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str or type(row[4]) == str:
                    continue

                # 1. Находим функциональную группу
                if not row[0] == None:
                    # Если длина кода равна 1, то дополним впереди 0: типа 01, 02 и т.д.
                    if len(str(row[0])) == 1:
                        code_fg = '0' + str(row[0])
                    else:
                        code_fg = str(row[0])
                    
                    zapis = funcgroup.objects.get(code = code_fg)
                    zapis.name_kaz = row[5]
                    zapis.save()
                    fg_id = zapis.id
                    continue  #Переходим к следующему циклу
                    #----------------------------------------------------------------- 


                # 2. Находим функциональную Подгруппу
                if not row[1] == None:
                    zapis = funcpodgroup.objects.get(code = row[1], _funcgroup = fg_id)
                    zapis.name_kaz = row[5]
                    zapis.save()
                    fpg_id = zapis.id
                    code_fpg = str(row[1])
                    continue  #Переходим к следующему циклу

                
                # Находим АБП
                if not row[2] == None:
                    try:
                        zapis = abp.objects.get(code = row[2])
                        zapis.name_kaz = row[5]
                        zapis.save()
                        abp_id = zapis.id
                        code_abp = str(row[2])
                    except:
                        print('abp code:' + str(row[2]))
                    continue  #Переходим к следующему циклу
                    #-----------------------------------------------------------------
                 



                # Находим программы
                if not row[3] == None:
                    if len(str(row[3])) == 1:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/00' + str(row[3])
                    elif len(str(row[3])) == 2:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/0' + str(row[3])
                    else:
                        code_pr = code_fg + '/' + code_fpg + '/' + code_abp + '/' + str(row[3])

                    try:
                        zapis = program.objects.get(code = code_pr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                        pr_id = zapis.id
                    except:
                        print('pr code:' + code_pr)
                    continue  #Переходим к следующему циклу
                    #-----------------------------------------------------------------



                # Находим Подпрограммы
                if not row[4] == None:
                    if len(str(row[4])) == 1:
                        code_ppr = code_pr + '/00' + str(row[4])
                    elif len(str(row[4])) == 2:
                        code_ppr = code_pr + '/0' + str(row[4])
                    else:
                        code_ppr = code_pr + '/' + str(row[4])

                    masscode = code_ppr.split('/')
                    codefkr = masscode[2] + '/' + masscode[3] + '/' + masscode[4]

                    try:
                        zapis = podprogram.objects.get(code = code_ppr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                        ppr_id = zapis.id
                    except:
                        print('ppr code:' + code_ppr)

                    try:
                        zapis = fkr.objects.get(code = codefkr)
                        zapis.name_kaz = row[5]
                        zapis.save()
                    except:
                        print('ppr code:' + codefkr)
                    continue  #Переходим к следующему циклу


            

        if lang == 'rus':
            try:
                with transaction.atomic():
                    funcgroup.objects.all().delete()
                    funcpodgroup.objects.all().delete()
                    program.objects.all().delete()
                    podprogram.objects.all().delete()
                    fkr.objects.all().delete()

                    # Читаем данные из ячеек в выбранном листе
                    for row in first_sheet.iter_rows(values_only=True):

                        # Пропускаем заголовки таблицы
                        if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str or type(row[4]) == str:
                            continue

                        # 1. Находим функциональную группу
                        if not row[0] == None:
                            # Если длина кода равна 1, то дополним впереди 0: типа 01, 02 и т.д.
                            if len(str(row[0])) == 1:
                                code = '0' + str(row[0])
                            else:
                                code = str(row[0])
                            # -----------------------------------------------------------------

                            # Если не существует, то создаем новую функциональную группу
                            funcgroupd, created = funcgroup.objects.get_or_create(code = code)
                            funcgroupd.name_rus = row[5]
                            if not existfg:
                                funcgroupd = funcgroup()
                                funcgroupd.code = code
                                funcgroupd.name_rus = row[5]
                                funcgroupd.save()
                                fg_id = funcgroupd.id
                                fg_code = code
                            # ---------------------------------------------- 

                        


                        # 2. Находим функциональную Подгруппу
                        if not row[1] == None:
                            code = str(row[1])
                            existfg = False
                            for itemfpg in obj_fpg:
                                if itemfpg[1] == code:
                                    existfg = True
                                    fpg_id = itemfpg[0]
                                    fpg_code = itemfpg[1]
                            if not existfg:
                                funcpodgroupd = funcpodgroup()
                                funcpodgroupd.code = code
                                funcpodgroupd.name_rus = row[5]
                                funcpodgroupd._funcgroup_id = fg_id
                                funcpodgroupd.save()
                                fpg_id = funcpodgroupd.id
                                fpg_code = funcpodgroupd.code

                        
                        # Находим АБП
                        if not row[2] == None:
                            code = str(row[2])
                            existfg = False
                            obj_abp = abp.objects.all().values_list()
                            for itemabp in obj_abp:
                                if itemabp[1] == code:
                                    existfg = True
                                    abp_id = itemabp[0]
                                    abp_code = itemabp[1]
                            if not existfg:
                                abpd = abp()                      
                                abpd.code = code
                                abpd.name_rus = row[5]
                                abpd.save()
                                abp_id = abpd.id
                                abp_code = abpd.code



                        # Находим программы
                        if not row[3] == None:
                            if len(str(row[3])) == 1:
                                code = fg_code + '/' + fpg_code + '/' + abp_code + '/00' + str(row[3])
                            elif len(str(row[3])) == 2:
                                code = fg_code + '/' + fpg_code + '/' + abp_code + '/0' + str(row[3])
                            else:
                                code = fg_code + '/' + fpg_code + '/' + abp_code + '/' + str(row[3])

        
                            existfg = False
                            for itempr in obj_prog:
                                if (itempr[1]) == code:
                                    existfg = True
                                    pr_id = itempr[0]
                                    pr_code = itempr[1]
                                    pr_name = itempr[3]
                            if not existfg:
                                progd = program()
                                progd.code =  code
                                progd.name_rus = row[5]
                                progd._funcgroup = fg_id
                                progd._funcpodgroup_id = fpg_id
                                progd._abp_id = abp_id
                                progd.save()
                                pr_id = progd.id
                                pr_code = progd.code
                                pr_name = progd.name_rus



                        # Находим Подпрограммы
                        if not row[4] == None:
                            if len(str(row[4])) == 1:
                                code = pr_code + '/00' + str(row[4])
                            elif len(str(row[4])) == 2:
                                code = pr_code + '/0' + str(row[4])
                            else:
                                code = pr_code + '/' + str(row[4])


                            existfg = False
                            for itempodpr in obj_podprog:
                                if (itempodpr[1]) == code:
                                    existfg = True
                                    ppr_id = itempodpr[0]
                                    ppr_code = itempodpr[1]

                            if not existfg:
                                podprogd = podprogram()
                                podprogd.code =  code
                                podprogd.name_rus = row[5]
                                podprogd._funcgroup = fg_id
                                podprogd._funcpodgroup = fpg_id
                                podprogd._abp = abp_id
                                podprogd._program_id = pr_id
                                podprogd.save()
                                ppr_id = podprogd.id
                                ppr_code = podprogd.code


                            # Создаем ФКР, имея уже все данные для создания
                            masscode = ppr_code.split('/')
                            codefkr = masscode[2] + '/' + masscode[3] + '/' + masscode[4]
                            existfg = False
                            for itemfkr in obj_fkr:
                                if (itemfkr[1]) == codefkr:
                                    existfg = True

                            if not existfg:
                                fkrnew = fkr()
                                fkrnew.code =  codefkr
                                fkrnew.name_rus = pr_name + '(' + row[5] + ')'
                                fkrnew._funcgroup = fg_id
                                fkrnew._funcpodgroup = fpg_id
                                fkrnew._abp = abp_id
                                fkrnew._program_id = pr_id
                                fkrnew._podprogram_id = ppr_id
                                fkrnew.save()
            except Exception as error:
                print(error)  
       
        # Закрываем книгу
        workbook.close()
    return 'Успешно'



def fkrreadxls(path='fkrnewkaz.xlsx'):
    
    if os.path.exists(path):  
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        first_sheet = workbook[sheet_names[0]]


        fg_id = 0
        fg_code = ''
        fg_name = ''
        fpg_id = 0
        fpg_code = ''
        fpg_name = ''
        abp_id = 0
        abp_code = ''
        abp_name = ''
        pr_id = 0
        pr_code = ''
        pr_name = ''
        ppr_id = 0
        ppr_code = ''
        ppr_name = ''

        lang = ''
        if first_sheet._cells[1,1].value == 'Функциональная группа':
            lang = 'rus'

        if first_sheet._cells[1,1].value == 'Функционалдық топ':
            lang = 'kaz'


        result = []
        count = 0
        if lang == 'rus':
            for row in first_sheet.iter_rows(values_only=True):
                count += 1
                if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str or type(row[4]) == str:
                    continue
 
                if not row[0] == None:
                    if len(str(row[0])) == 1:
                        fg_code = '0' + str(row[0])
                    else:
                        fg_code = str(row[0])
                    fg_name = row[5]
                    
                if not row[1] == None and not fg_code == '':
                    fpg_code = str(row[1])
                    fpg_name = row[5]


                if not row[2] == None and not fpg_code == '':
                    abp_code = str(row[2])
                    abp_name = row[5]


                if not row[3] == None and not abp_code == '':
                    if len(str(row[3])) == 1:
                        pr_code = '00' + str(row[3])
                    elif len(str(row[3])) == 2:
                        pr_code = '0' + str(row[3])
                    else:
                        pr_code = str(row[3])
                    pr_name = row[5]

                    if not first_sheet._cells[count+1, 4].value == None or \
                            not first_sheet._cells[count+1, 3].value == None or \
                                not first_sheet._cells[count+1, 2].value == None or \
                                    not first_sheet._cells[count+1, 1].value == None: \
                                        result.append(
                                                {
                                                    "fg_code":fg_code,
                                                    "fpg_code":fpg_code,
                                                    "abp_code":abp_code,
                                                    "pr_code":pr_code,
                                                    "ppr_code":'000',
                                                    "fg_name":fg_name,
                                                    "fpg_name":fpg_name,
                                                    "abp_name":abp_name,
                                                    "pr_name":pr_name,
                                                    "ppr_name":pr_name,
                                                }
                                            )


                if not row[4] == None and not pr_code == '':
                    if len(str(row[4])) == 1:
                        ppr_code = '00' + str(row[4])
                    elif len(str(row[4])) == 2:
                        ppr_code = '0' + str(row[4])
                    else:
                        ppr_code = str(row[4])
                    ppr_name = row[5]
                    result.append({
                                    "fg_code":fg_code,
                                    "fpg_code":fpg_code,
                                    "abp_code":abp_code,
                                    "pr_code":pr_code,
                                    "ppr_code":ppr_code,
                                    "fg_name":fg_name,
                                    "fpg_name":fpg_name,
                                    "abp_name":abp_name,
                                    "pr_name":pr_name,
                                    "ppr_name":ppr_name,
                                })
        # mas_code = []
        # for row in result:
        #     code = row['fg_code'] + '/' +row['fpg_code'] + '/' +row['abp_code'] + '/' +row['pr_code'] + '/' +row['ppr_code']
        #     mas_code.append(code)

        query = """SELECT code FROM public.dirs_podprogram"""
        with connection.cursor() as cursor:
            cursor.execute(query)
            res_bd = cursor.fetchall()

        cnt = 0
        for row in result:
            fullcode = row['fg_code'] + '/' +row['fpg_code'] + '/' + row['abp_code'] + '/' +row['pr_code'] + '/' +row['ppr_code']
            exist = False
            for j in res_bd:
                if j[0]==fullcode:
                    exist = True
                    break

            if not exist:
                cnt += 1
                fg_obj, fg_create = funcgroup.objects.get_or_create(code = row['fg_code'])
                if fg_create:
                    fg_obj.name_rus = row['fg_name']
                    fg_obj.save()


                fpg_obj, fpg_create = funcpodgroup.objects.get_or_create(code = row['fpg_code'], _funcgroup_id = fg_obj.id)
                if fpg_create:
                    fpg_obj.name_rus = row['fpg_name']
                    fpg_obj.save()
                

                abp_obj, abp_create = abp.objects.get_or_create(code = row['abp_code'])
                if abp_create:
                    abp_obj.name_rus = row['abp_name']
                    abp_obj.save()


                pr_obj, pr_create = program.objects.get_or_create(code = row['fg_code'] + '/' +row['fpg_code'] + '/' +row['abp_code'] + '/' +row['pr_code'], _abp_id = abp_obj.id, _funcpodgroup_id = fpg_obj.id)
                if pr_create:
                    pr_obj.name_rus = row['pr_name']
                    pr_obj._funcgroup = fg_obj.id
                    pr_obj._abp_id = abp_obj.id
                    pr_obj._funcpodgroup_id = fpg_obj.id
                    pr_obj.save()


                ppr_obj, ppr_create = podprogram.objects.get_or_create(code = fullcode, _program_id = pr_obj.id)
                if ppr_create:
                    ppr_obj.name_rus = row['ppr_name']
                    ppr_obj._funcgroup = fg_obj.id
                    ppr_obj._abp = abp_obj.id
                    ppr_obj._funcpodgroup = fpg_obj.id
                    ppr_obj._program_id = pr_obj.id
                    ppr_obj.save()

                fkr_obj, fkr_create = fkr.objects.get_or_create(code = row['abp_code'] + '/' +row['pr_code'] + '/' +row['ppr_code'], _program_id = pr_obj.id, _podprogram_id = ppr_obj.id)
                if fkr_create:
                    fkr_obj._funcgroup = fg_obj.id
                    fkr_obj._funcpodgroup = fpg_obj.id
                    fkr_obj._abp = abp_obj.id
                    fkr_obj.name_rus = row['ppr_name']
                    fkr_obj.save()

                print(fullcode)
  

   
        print(cnt)
        # Закрываем книгу
        workbook.close()
    return 'Успешно'



def ekrreadxls(path='ekrnewkaz.xlsx'):  
    if os.path.exists(path):  
        obj_ekr = spec_exp.objects.all().values_list()

 
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        first_sheet = workbook[sheet_names[0]]

        lang = ''
        if first_sheet._cells[1,1].value == 'Подкласс':
            lang = 'rus'

        if first_sheet._cells[1,1].value == 'Функционалдық топ':
            lang = 'kaz'

        if lang == '':
            return 'error excel'
       

        if lang == 'kaz':
            # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[1]) == str:
                    continue

                if not row[1] == None:
                    code = str(row[1])
                    
                    zapis = spec_exp.objects.get(code = code)
                    zapis.name_kaz = row[2]
                    zapis.save()
                    #----------------------------------------------------------------- 


            

        if lang == 'rus':
             # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[1]) == str:
                    continue

                if not row[1] == None:
                    code = str(row[1])
    
                    # Ищем, есть ли в базе функ группа с таким кодом
                    existfg = False
                    for itemfg in obj_ekr:
                        if itemfg[1] == code:
                            existfg = True

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = spec_exp()
                        newexp.code = code
                        newexp.name_rus = row[2]
                        newexp.save()
                    # ---------------------------------------------- 
                         
        # Закрываем книгу
        workbook.close()
    return 'Успешно'




def inc_dir_import_xls(path='dohod_codes.xlsx'):  
    if os.path.exists(path):  
        obj_cat = category_income.objects.all().values_list()
        obj_clas = class_income.objects.all().values_list()
        obj_podcl = podclass_income.objects.all().values_list()
        obj_spec = spec_income.objects.all().values_list()
        obj_classific = classification_income.objects.all().values_list()

        cat_id = 0
        clas_id = 0
        podcl_id = 0
        spec_id = 0

        cat_code = ''
        clas_code = ''
        podcl_code = ''
        spec_code = ''

 
        workbook = load_workbook(path)
        sheet_names = workbook.sheetnames
        first_sheet = workbook[sheet_names[0]]

        lang = ''
        if first_sheet._cells[1,1].value == 'Категория':
            lang = 'rus'

        if first_sheet._cells[1,1].value == 'Функционалдық топ':
            lang = 'kaz'

        if lang == '':
            return 'error excel'
       

        if lang == 'kaz':
            # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[0]) == str:
                    continue

                if not row[0] == None:
                    code = str(row[0])
                    
                    zapis = category_income.objects.get(code = code)
                    zapis.name_kaz = row[4]
                    zapis.save()
                    #----------------------------------------------------------------- 


            

        if lang == 'rus':
             # Читаем данные из ячеек в выбранном листе
            for row in first_sheet.iter_rows(values_only=True):

                # Пропускаем заголовки таблицы
                if type(row[0]) == str or type(row[1]) == str or type(row[2]) == str or type(row[3]) == str:
                    continue

                if not row[0] == None :
                    code = str(row[0])
                    cat_code = code

                    # Ищем, есть ли в базе
                    existfg = False
                    for itemfg in obj_cat:
                        if itemfg[1] == code:
                            existfg = True
                            cat_id = itemfg[0]

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = category_income()
                        newexp.code = code
                        newexp.name_rus = row[4]
                        newexp.save()
                        cat_id = newexp.id
                    # ---------------------------------------------- 


                if not row[1] == None :
                    code = str(row[1])
                    if len(code) == 1:
                        code = '0' + code
                    clas_code = cat_code + '/' + code

    
                    # Ищем, есть ли в базе функ группа с таким кодом
                    existfg = False
                    for itemfg in obj_clas:
                        if itemfg[1] == clas_code:
                            existfg = True
                            clas_id = itemfg[0]

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = class_income()
                        newexp._category_id = cat_id
                        newexp.code = clas_code
                        newexp.name_rus = row[4]
                        newexp.save()
                        clas_id = newexp.id
                    # ----------------------------------------------

                if not row[2] == None :
                    code = str(row[2])
                    podcl_code = clas_code + '/' + code
    
                    # Ищем, есть ли в базе функ группа с таким кодом
                    existfg = False
                    for itemfg in obj_podcl:
                        if itemfg[1] == podcl_code:
                            existfg = True
                            podcl_id = itemfg[0]

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = podclass_income()
                        newexp._classs_id = clas_id
                        newexp.code = podcl_code
                        newexp.name_rus = row[4]
                        newexp.save()
                        podcl_id = newexp.id
                    # ----------------------------------------------


                if not row[3] == None :
                    code = str(row[3])
                    if len(code) == 1:
                        code = '0' + code
                    spec_code = podcl_code + '/' + code
    
                    # Ищем, есть ли в базе функ группа с таким кодом
                    existfg = False
                    for itemfg in obj_spec:
                        if itemfg[1] == spec_code:
                            existfg = True
                            spec_id = itemfg[0]

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = spec_income()
                        newexp._podclass_id = podcl_id
                        newexp.code = spec_code
                        newexp.name_rus = row[4]
                        newexp.save()
                        spec_id = newexp.id
                    # ----------------------------------------------



                    # Классификация по поступлениям
                    existfg = False
                    for itemfg in obj_classific:
                        if itemfg[1] == spec_code:
                            existfg = True
                            spec_id = itemfg[0]

                    # Если не существует, то создаем новую
                    if not existfg:
                        newexp = classification_income()
                        newexp._category_id = cat_id
                        newexp._classs_id = clas_id
                        newexp._podclass_id = podcl_id
                        newexp._spec_id = spec_id
                        newexp.code = spec_code
                        newexp.name_rus = row[4]
                        newexp.save()
                        spec_id = newexp.id
                    # ----------------------------------------------
                         
        # Закрываем книгу
        workbook.close()
    return 'Успешно'





