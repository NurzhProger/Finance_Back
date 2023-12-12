from datetime import datetime
from django.db import connection, transaction
from .models import *
from openpyxl import load_workbook
import os
from docs.serializer import *
# from PyPDF2 import PdfReader


def getqsetlist(modelname, userobj, sortcolumn, tip=None):
    not_adm = userobj.groups.filter(name='fulldata1').count()==0
    id_org = userobj.profile._organization_id
    if not_adm:
        if tip=='select':
            ids = [id_org]
            orgs = parent_organizations.objects.filter(_parent_id = id_org).values('_organization_id')
            for i in orgs:
                ids.append(i['_organization_id'])
            queryset = modelname.objects.filter(_organization_id__in = ids).order_by(sortcolumn)
        else:
            queryset = modelname.objects.filter(_organization_id = id_org).order_by(sortcolumn)
    else:
        queryset = modelname.objects.order_by(sortcolumn)
    return queryset


def fkrreadxls(path='fkr.xls'):
    listnum = ('0', '1', '2', '3', '4', '5', '6', '7', '8', '9')

    if os.path.exists(path):
        # Загружаем книгу Excel
        workbook = load_workbook(path)
        # Получаем список названий листов в книге
        sheet_names = workbook.sheetnames
        # Выбираем первый лист
        first_sheet = workbook[sheet_names[0]]

        # Читаем данные из ячеек в выбранном листе
        rowcount = 0
        for row in first_sheet.iter_rows(values_only=True):
            str_0 = row[0]
            str_1 = row[1]

        # Закрываем книгу
        workbook.close()
    return True


def object_svod_get(id_doc):
    queryset = type_izm_doc.objects.all()
    serialtype = typedocSerializer(queryset, many=True)

    jsondata = {
        "doc": {
            "id": 0,
            "nom": "",
            "_date": datetime.now().strftime("%d.%m.%Y %H:%M:%S"),
            "deleted": False,
            "_organization": {
                "id": 0,
                "name_rus": ""
            },
            "_type_izm_doc": {
                "id": 0,
                "name_rus": ""
            }
        },
        "payments": [],
        "obligats": [],
        "docs_izm": []
    }
    


    querydoc = f"""with doc as (SELECT * FROM public.docs_svod_exp WHERE id = {id_doc} ),
                        typeizm as (select * from dirs_type_izm_doc),
                        org as (select id, bin, name_rus from public.dirs_organization
                                where id in (select _organization_id from doc)
                                )
                                
                    select doc.id, doc.nom, to_char(doc._date, 'dd.mm.yyyy hh:mm:ss') as _date, doc.deleted, org.id as _organization_id, org.name_rus as _organization_name, typeizm.id as _type_izm_doc_id, typeizm.name_rus as _type_izm_doc_name 
                    from org, doc
                    left join typeizm
                    on doc._type_izm_doc_id = typeizm.id"""

    querydoctbl = f"""with doc as (SELECT * FROM docs_svod_exp_tbl
                                WHERE _svod_exp_id = {id_doc}
                                ),
                        izm_docs as (select 'izm' as tipdoc, id, deleted, nom, _date,_organization_id  from docs_izm_exp
                                where id in (select _izm_exp_id from doc)
                                ),
						svod_docs as (select 'svod' as tipdoc, id, deleted, nom, _date, _organization_id  from docs_svod_exp
                                where id in (select _svod_exp1_id from doc)
                                ),
                        org as (select id, bin, name_rus from public.dirs_organization
                                where id in (select _organization_id from doc)
                                ),
						uniondocs as (
							select * from izm_docs
							union all
							select * from svod_docs
						)              
                        select doc.id, uniondocs.tipdoc, uniondocs.id as izm_id, uniondocs.deleted, uniondocs.nom,  to_char(uniondocs._date, 'dd.mm.yyyy hh:mm:ss') as _date, 
                                org.id as _organization_id, org.name_rus as _organization_name
                        from doc
                        left join uniondocs
                            on doc._izm_exp_id = uniondocs.id or doc._svod_exp1_id = uniondocs.id
                        left join org
                            on uniondocs._organization_id = org.id"""

    querypay = f"""with pay as (SELECT * FROM public.docs_izm_exp_pay
                                WHERE _izm_exp_id in %s),
                        fkr as (select id as _fkr_id, code as _fkr_code, name_rus as _fkr_name from public.dirs_fkr
                                WHERE id in (select _fkr_id from pay)),
                        spec as (select id as _spec_id, code as _spec_code, name_rus as _spec_name from public.dirs_spec_exp
                                WHERE id in (select _spec_id from pay)),
                itog as (select  fkr._fkr_id, fkr._fkr_code, fkr._fkr_name, 
							spec._spec_id, spec._spec_code, spec._spec_name, 
							pay.sm1+pay.sm2+pay.sm3+pay.sm4+pay.sm5+pay.sm6+pay.sm7+pay.sm8+pay.sm9+pay.sm10+pay.sm11+pay.sm12 as god, 
							pay.sm1, pay.sm2, pay.sm3, pay.sm4, pay.sm5, pay.sm6, pay.sm7, pay.sm8, pay.sm9, pay.sm10, pay.sm11, pay.sm12 from pay
                    left join fkr
                    on pay._fkr_id = fkr._fkr_id
                    left join spec
                    on pay._spec_id = spec._spec_id
                    order by fkr._fkr_code, spec._spec_code)
	
                select _fkr_id, max(_fkr_code) as _fkr_code, max(_fkr_name) as _fkr_name, _spec_id, max(_spec_code) as _spec_code, max(_spec_name) as _spec_name, 
                sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12
                from itog
                group by _fkr_id, _spec_id"""

    queryobl = f"""with pay as (SELECT * FROM public.docs_izm_exp_obl
                                WHERE _izm_exp_id in %s),
                        fkr as (select id as _fkr_id, code as _fkr_code, name_rus as _fkr_name from public.dirs_fkr
                                WHERE id in (select _fkr_id from pay)),
                        spec as (select id as _spec_id, code as _spec_code, name_rus as _spec_name from public.dirs_spec_exp
                                WHERE id in (select _spec_id from pay)),
                itog as (select  fkr._fkr_id, fkr._fkr_code, fkr._fkr_name, 
							spec._spec_id, spec._spec_code, spec._spec_name, 
							pay.sm1+pay.sm2+pay.sm3+pay.sm4+pay.sm5+pay.sm6+pay.sm7+pay.sm8+pay.sm9+pay.sm10+pay.sm11+pay.sm12 as god, 
							pay.sm1, pay.sm2, pay.sm3, pay.sm4, pay.sm5, pay.sm6, pay.sm7, pay.sm8, pay.sm9, pay.sm10, pay.sm11, pay.sm12 from pay
                    left join fkr
                    on pay._fkr_id = fkr._fkr_id
                    left join spec
                    on pay._spec_id = spec._spec_id
                    order by fkr._fkr_code, spec._spec_code)
	
                select _fkr_id, max(_fkr_code) as _fkr_code, max(_fkr_name) as _fkr_name, _spec_id, max(_spec_code) as _spec_code, max(_spec_name) as _spec_name, 
                sum(sm1) as sm1, sum(sm2) as sm2, sum(sm3) as sm3, sum(sm4) as sm4, sum(sm5) as sm5, sum(sm6) as sm6, sum(sm7) as sm7, sum(sm8) as sm8, sum(sm9) as sm9, sum(sm10) as sm10, sum(sm11) as sm11, sum(sm12) as sm12
                from itog
                group by _fkr_id, _spec_id"""


    with connection.cursor() as cursor:
        cursor.execute(querydoc)
        columns = [col[0] for col in cursor.description]
        result = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

        cursor.execute(querydoctbl)
        columns = [col[0] for col in cursor.description]
        resulttbl = [dict(zip(columns, row))
                  for row in cursor.fetchall()]

        qset = reg_svod_exp.objects.filter(_svod_exp_id = id_doc)
        izm_id_mass = []
        izm_id_mass.append(0)
        for itm in qset:
            izm_id_mass.append(itm._izm_exp_id)


        cursor.execute(querypay, (tuple(izm_id_mass),))
        columns = [col[0] for col in cursor.description]
        resultpay = [dict(zip(columns, row))
                     for row in cursor.fetchall()]

        cursor.execute(queryobl, (tuple(izm_id_mass),))
        columns = [col[0] for col in cursor.description]
        resultobl = [dict(zip(columns, row))
                     for row in cursor.fetchall()]

    jsondata['doc']['id'] = result[0]['id']
    jsondata['doc']['_date'] = result[0]['_date']
    jsondata['doc']['nom'] = result[0]['nom']
    jsondata['doc']['_organization']['id'] = result[0]['_organization_id']
    jsondata['doc']['_organization']['name_rus'] = result[0]['_organization_name']
    jsondata['doc']['_type_izm_doc']['id'] = result[0]['_type_izm_doc_id']
    jsondata['doc']['_type_izm_doc']['name_rus'] = result[0]['_type_izm_doc_name']
    jsondata['doc']['deleted'] = result[0]['deleted']
    jsondata['payments'] = resultpay
    jsondata['obligats'] = resultobl
    jsondata['docs_izm'] = resulttbl

    jsondata["typesdoc"] = serialtype.data

    return jsondata



def simplepagination(offset, limit, request, result):
    current_url = request.build_absolute_uri()
    if offset>=limit:
        previous = current_url + "?limit=" + str(limit) +  "&offset=" + str(int(offset) - int(limit))
        next = current_url + "?limit=" + str(limit) +  "&offset=" + str(int(offset) + int(limit))
    else:
        previous = None
        next = current_url + "?limit=" + str(limit) +  "&offset=" + str(int(offset) + int(limit))

    if len(result)<int(limit):
        next = None
        offset_prev = max(int(offset) - int(limit), 0)
        if offset_prev==0:
            previous = None
        else:
            previous = current_url + "?limit=" + str(limit) +  "&offset=" + str(offset_prev)
    
    if len(result)>0:
        count = result[0]['count']
    else:
        count = 0

    response = {
        "count": count,
        "next": next,
        "previous": previous,
        "results": result
        }
    return response











